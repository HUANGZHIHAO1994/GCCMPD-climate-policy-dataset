import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Border, Side, PatternFill, Font
import os
import sys

print("================== {} ==================".format(os.path.basename(sys.argv[0])))

# blue #00B0F0 orange #FFC000 green #92D050 red #FF0000 brown #E26B0A
# 'FF00B0F0', 'FFFFC000', 'FF92D050', 'FFFF0000', 'FFC00000'
# df_1: Index to color
filename = "previous_bm25_dedup_result.xlsx"
workbook = openpyxl.load_workbook(filename)
worksheet = workbook["all_policies_dedup"]
rows, cols = worksheet.max_row, worksheet.max_column
# print(rows)
# print(cols)
colors_dict = dict()
colors = []

for i in range(2, rows + 1):
    ce = worksheet.cell(row=i, column=1)
    # print(ce)
    fill = ce.fill
    font = ce.font
    # print(font.color.rgb)
    if font.color.rgb in ['FF00B0F0', 'FFFFC000', 'FF92D050', 'FFFF0000', 'FFC00000', '0000B0F0', '00FFC000',
                          '0092D050', '00FF0000', '00C00000']:
        colors_dict[ce.value] = font.color.rgb
    colors.append(font.color.rgb)
# print(set(colors))
# print(colors_dict)
pd.set_option('display.max_columns', 4)

df_1_all_policies_dedup = pd.read_excel("previous_bm25_dedup_result.xlsx", sheet_name="all_policies_dedup")
df_1_dup = pd.read_excel("previous_bm25_dedup_result.xlsx", sheet_name="dup")

df_2 = pd.read_excel("bm25_result.xlsx")
df_2['Year_amend'] = df_2['Year']
df_2['Jurisdiction_standard_amend'] = df_2['Jurisdiction_standard']

# df_1_all_policies_dedup.rename(columns={"Index": "A"}, inplace=True)
# df_1_all_policies_dedup.rename(columns={"Index": "A"}, inplace=True)
# df_1_dup.rename(columns={"Unnamed: 0": "A"}, inplace=True)
df_2.rename(columns={"Unnamed: 0": "A"}, inplace=True)

df_2 = df_2[['A', 'Policy', 'Short_name', 'ISO_code', 'Year', 'Year_amend',
             'law or strategy', "Policy Type", 'IPCC_Region', 'WB_Region', 'sector', 'subsector',
             'Instrument', 'Sector-Instrument', 'objective', 'subobjective',
             'Long_name', 'Income_Group', 'Annex', 'Policy_Content', 'Status',
             'Jurisdiction', 'db_source', 'Source', 'Jurisdiction_standard', 'Jurisdiction_standard_amend',
             'group_policy_number', 'bm25_score_first', 'bm25_score_second',
             'bm25_policy_first', 'bm25_index_first', 'bm25_policy_second', 'bm25_index_second']]
# print(df_1_all_policies_dedup.info())
# print(df_1_dup.info())
# print(df_2.info())


# data matched in all_policies_dedup sheet
df_2_match = []

# Unmatched data
df_3_match = []

# data matched in dup sheet
df_2_dup_match = []

# match index，df_1 is the key
df_1_df_2_match_dict = dict()

for num, row in df_2.iterrows():
    ISO_code = row["ISO_code"]
    Year = row["Year"]
    Jurisdiction_standard = row["Jurisdiction_standard"]
    Policy = row["Policy"]
    db_source = row["db_source"]

    df_1_all_policies_dedup_match = df_1_all_policies_dedup.query(
        "ISO_code==@ISO_code & Year==@Year & Jurisdiction_standard==@Jurisdiction_standard & Policy==@Policy & db_source==@db_source")
    df_1_dup_match = df_1_dup.query(
        "ISO_code==@ISO_code & Year==@Year & Jurisdiction_standard==@Jurisdiction_standard & Policy==@Policy & db_source==@db_source")

    # group_policy_number	bm25_score_first	bm25_score_second	bm25_policy_first	bm25_index_first	bm25_policy_second	bm25_index_second
    # 将匹配到的数据 对上方字段进行修改（修改字段较少）然后将修改后的数据添加到df_2_after_match列表里
    # 最后再将df_2_after_match替换原bm25_result.xlsx的内容 这样就相当于修改了除上方字段的其他字段
    if len(df_1_all_policies_dedup_match) > 0:
        if len(df_1_all_policies_dedup_match) > 1:
            print("***********all_policies_dedup_match***********")
            print(df_1_all_policies_dedup_match)
            # Delete the matching data in the original data
            for inner_num, inner_row in df_1_all_policies_dedup_match.iterrows():
                df_1_all_policies_dedup_match_A = inner_row["A"]
                df_1_all_policies_dedup = df_1_all_policies_dedup.query("A != @df_1_all_policies_dedup_match_A")
        else:
            df_1_all_policies_dedup_match_A = df_1_all_policies_dedup_match["A"].tolist()[0]
            df_1_all_policies_dedup = df_1_all_policies_dedup.query("A != @df_1_all_policies_dedup_match_A")

        # If there are multiple matches, take the first one
        df_1_all_policies_dedup_match = df_1_all_policies_dedup_match.iloc[0, :]
        df_1_df_2_match_dict[df_1_all_policies_dedup_match["A"]] = row["A"]
        df_1_all_policies_dedup_match["A"] = row["A"]
        df_1_all_policies_dedup_match["group_policy_number"] = row["group_policy_number"]
        df_1_all_policies_dedup_match["bm25_score_first"] = row["bm25_score_first"]
        df_1_all_policies_dedup_match["bm25_score_second"] = row["bm25_score_second"]
        df_1_all_policies_dedup_match["bm25_policy_first"] = row["bm25_policy_first"]
        df_1_all_policies_dedup_match["bm25_index_first"] = row["bm25_index_first"]
        df_1_all_policies_dedup_match["bm25_policy_second"] = row["bm25_policy_second"]
        df_1_all_policies_dedup_match["bm25_index_second"] = row["bm25_index_second"]
        df_1_all_policies_dedup_match["Policy Type"] = row["Policy Type"]
        df_2_match.append(df_1_all_policies_dedup_match)
    elif len(df_1_dup_match) > 0:
        if len(df_1_dup_match) > 1:
            print("***********dup_match***********")
            print(df_1_dup_match)
            # Delete the matching data in the original data
            for inner_num, inner_row in df_1_dup_match.iterrows():
                df_1_dup_match_A = inner_row["A"]
                df_1_dup = df_1_dup.query("A != @df_1_dup_match_A")
        else:
            df_1_dup_match_A = df_1_dup_match["A"].tolist()[0]
            df_1_dup = df_1_dup.query("A != @df_1_dup_match_A")
        # If there are multiple matches, take the first one
        df_1_dup_match = df_1_dup_match.iloc[0, :]

        df_1_dup_match["A"] = row["A"]
        df_1_df_2_match_dict[df_1_dup_match["A"]] = row["A"]

        df_1_dup_match["law or strategy"] = row["law or strategy"]
        df_1_dup_match["Policy Type"] = row["Policy Type"]
        df_1_dup_match["sector"] = row["sector"]
        df_1_dup_match["subsector"] = row["subsector"]
        df_1_dup_match["Instrument"] = row["Instrument"]
        df_1_dup_match["Sector-Instrument"] = row["Sector-Instrument"]
        df_1_dup_match["objective"] = row["objective"]
        df_1_dup_match["subobjective"] = row["subobjective"]

        df_1_dup_match["group_policy_number"] = row["group_policy_number"]
        df_1_dup_match["bm25_score_first"] = row["bm25_score_first"]
        df_1_dup_match["bm25_score_second"] = row["bm25_score_second"]
        df_1_dup_match["bm25_policy_first"] = row["bm25_policy_first"]
        df_1_dup_match["bm25_index_first"] = row["bm25_index_first"]
        df_1_dup_match["bm25_policy_second"] = row["bm25_policy_second"]
        df_1_dup_match["bm25_index_second"] = row["bm25_index_second"]
        # print(df_1_dup_match)
        df_2_dup_match.append(df_1_dup_match)
    else:
        # print("***********not_match***********")
        # print(row.tolist())
        df_3_match.append(row)

# to dataFrame
df_2_match_df = pd.DataFrame(df_2_match)
df_2_match_df \
    = df_2_match_df[['A', 'Policy', 'Short_name', 'ISO_code', 'Year', 'Year_amend',
                     'law or strategy', "Policy Type", 'IPCC_Region', 'WB_Region', 'sector', 'subsector', 'Instrument',
                     'Sector-Instrument', 'objective', 'subobjective', 'Long_name', 'Income_Group', 'Annex',
                     'Policy_Content', 'Status', 'Jurisdiction', 'db_source', 'Source', 'Jurisdiction_standard',
                     'Jurisdiction_standard_amend',
                     'group_policy_number', 'bm25_score_first', 'bm25_score_second',
                     'bm25_policy_first', 'bm25_index_first', 'bm25_policy_second', 'bm25_index_second']]

# print(df_3_match)
if df_3_match:
    df_3_match_df = pd.DataFrame(df_3_match)
    # print(df_3_match_df.info())
    df_3_match_df \
        = df_3_match_df[['A', 'Policy', 'Short_name', 'ISO_code', 'Year', 'Year_amend',
                         'law or strategy', "Policy Type", 'IPCC_Region', 'WB_Region', 'sector', 'subsector',
                         'Instrument',
                         'Sector-Instrument', 'objective', 'subobjective', 'Long_name', 'Income_Group', 'Annex',
                         'Policy_Content', 'Status', 'Jurisdiction', 'db_source', 'Source', 'Jurisdiction_standard',
                         'Jurisdiction_standard_amend',
                         'group_policy_number', 'bm25_score_first', 'bm25_score_second',
                         'bm25_policy_first', 'bm25_index_first', 'bm25_policy_second', 'bm25_index_second']]
    result_df = pd.concat([df_2_match_df, df_3_match_df])
else:
    result_df = df_2_match_df

result_df = result_df.sort_values(by=['group_policy_number', 'bm25_score_first', 'Policy'],
                                  ascending=[False, False, False])

df_2_dup_match_df = pd.DataFrame(df_2_dup_match)
df_2_dup_match_df \
    = df_2_dup_match_df[['A', 'Policy', 'Short_name', 'ISO_code', 'Year', 'Year_amend',
                         'law or strategy', "Policy Type", 'IPCC_Region', 'WB_Region', 'sector', 'subsector',
                         'Instrument',
                         'Sector-Instrument', 'objective', 'subobjective', 'Long_name', 'Income_Group', 'Annex',
                         'Policy_Content', 'Status', 'Jurisdiction', 'db_source', 'Source', 'Jurisdiction_standard',
                         'Jurisdiction_standard_amend',
                         'group_policy_number', 'bm25_score_first', 'bm25_score_second',
                         'bm25_policy_first', 'bm25_index_first', 'bm25_policy_second', 'bm25_index_second',
                         'dup_source', 'dup_index']]

df_2_dup_match_df['dup_index'] = df_2_dup_match_df['dup_index'].apply(lambda x: df_1_df_2_match_dict.get(x, ''))
df_2_dup_match_df = df_2_dup_match_df.sort_values(by=['dup_index', 'dup_source'], ascending=[False, False])

# output
target_and_ndc = pd.read_excel("previous_bm25_result.xlsx", sheet_name="Target and NDC", engine='openpyxl')
else_target = pd.read_excel("previous_bm25_result.xlsx", sheet_name="else Target", engine='openpyxl')

with pd.ExcelWriter('result_new_color.xlsx') as writer:
    df_2.to_excel(writer, index=False, sheet_name="raw_data")
    result_df.to_excel(writer, index=False, sheet_name="all_policies_dedup")
    target_and_ndc.to_excel(writer, index=False, sheet_name="Target and NDC")
    else_target.to_excel(writer, index=False, sheet_name="else Target")
    df_2_dup_match_df.to_excel(writer, index=False, sheet_name="dup")

with pd.ExcelWriter('result_new_not_match.xlsx') as writer:
    df_1_all_policies_dedup.to_excel(writer, index=False, sheet_name="all_policies_dedup")
    df_1_dup.to_excel(writer, index=False, sheet_name="dup")

# color: df_2_match_df (Color according to original document)，df_3_match_df (Purple: 800080)
df2_colors_dict = {value: colors_dict[key] for key, value in df_1_df_2_match_dict.items() if colors_dict.get(key, None)}
# print(df2_colors_dict)
filename = "result_new_color.xlsx"
workbook = openpyxl.load_workbook(filename)
worksheet = workbook["all_policies_dedup"]
rows, cols = worksheet.max_row, worksheet.max_column

for i in range(2, len(df_2_match_df) + 2):
    ce = worksheet.cell(row=i, column=1)
    if ce.value in df2_colors_dict:
        for j in range(1, cols + 1):
            ce_temp = worksheet.cell(row=i, column=j)
            ce_temp.font = Font(color=df2_colors_dict[ce.value][2:])
        # row = worksheet.row_dimensions[i]
        # # print(row.font)
        # row.font = Font(color=colors_dict[ce.value][2:])
for i in range(len(df_2_match_df) + 2, rows + 1):
    for j in range(1, cols + 1):
        ce_temp = worksheet.cell(row=i, column=j)
        ce_temp.font = Font(color='800080')

workbook.save("result_new_color.xlsx")
