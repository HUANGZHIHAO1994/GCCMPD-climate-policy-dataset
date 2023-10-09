import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Border, Side, PatternFill, Font
import os
import sys

print("================== {} ==================".format(os.path.basename(sys.argv[0])))


def get_data():
    df_all_policies_dedup = pd.read_excel("previous_bm25_result.xlsx", sheet_name="all_policies_dedup",
                                          engine='openpyxl')
    df_dup = pd.read_excel("previous_bm25_result.xlsx", sheet_name="dup", engine='openpyxl')
    # print(df_dup.info())
    return df_all_policies_dedup, df_dup


def data_process(df_all_policies_dedup, df_dup, colors_dict):
    df_all_policies_dedup.fillna('', inplace=True)
    df_dup.fillna('', inplace=True)
    # print(len(df_all_policies_dedup))
    df_all_policies_dedup = df_all_policies_dedup.sort_values(by=['group_policy_number', 'Policy'],
                                                              ascending=[False, False])
    df_all_policies_dedup = df_all_policies_dedup.drop_duplicates(
        subset=["ISO_code", "Year", "Jurisdiction_standard", "Policy", "db_source"], keep="first")
    # print(len(df_all_policies_dedup))

    # print(len(df_dup))
    df_dup = df_dup.drop_duplicates(subset=["ISO_code", "Year", "Jurisdiction_standard", "Policy", "db_source"],
                                    keep="first")
    # print(len(df_dup))

    with pd.ExcelWriter("previous_bm25_dedup_result.xlsx", engine='xlsxwriter') as writer:
        df_all_policies_dedup.to_excel(writer, index=False, sheet_name="all_policies_dedup")
        df_dup.to_excel(writer, index=False, sheet_name="dup")

    filename = "previous_bm25_dedup_result.xlsx"
    workbook = openpyxl.load_workbook(filename)
    worksheet = workbook["all_policies_dedup"]
    rows, cols = worksheet.max_row, worksheet.max_column

    # col = worksheet.column_dimensions['A']
    # col.font = Font(bold=True)
    # d4 = worksheet['D5']
    # ft = Font(color="00B0F0")
    # d4.font = ft
    # d5 = worksheet['D4']
    # ft = Font(color="92D050")
    # d5.font = ft
    # d6 = worksheet['D7']
    # ft = Font(color="C00000")
    # d6.font = ft

    for i in range(2, rows + 1):
        ce = worksheet.cell(row=i, column=1)
        if ce.value in colors_dict:
            for j in range(1, cols + 1):
                ce_temp = worksheet.cell(row=i, column=j)
                ce_temp.font = Font(color=colors_dict[ce.value][2:])
            # row = worksheet.row_dimensions[i]
            # # print(row.font)
            # row.font = Font(color=colors_dict[ce.value][2:])

    workbook.save("previous_bm25_dedup_result.xlsx")


if __name__ == '__main__':

    # blue #00B0F0 orange #FFC000 green #92D050 red #FF0000 brown #E26B0A
    # 'FF00B0F0', 'FFFFC000', 'FF92D050', 'FFFF0000', 'FFC00000'

    # read excel
    filename = "previous_bm25_result.xlsx"
    workbook = openpyxl.load_workbook(filename)

    # read sheet
    worksheet = workbook["all_policies_dedup"]
    rows, cols = worksheet.max_row, worksheet.max_column
    # print(rows)
    # print(cols)
    colors_dict = dict()

    for i in range(2, rows + 1):
        ce = worksheet.cell(row=i, column=1)
        # print(ce)
        fill = ce.fill
        font = ce.font
        # print(font.color)
        if font.color.rgb in \
                ['FF00B0F0', 'FFFFC000', 'FF92D050', 'FFFF0000', 'FFC00000', '0000B0F0', '00FFC000',
                 '0092D050', '00FF0000', '00C00000']:
            colors_dict[ce.value] = font.color.rgb
    # print(set(colors))
    all_policies_dedup_data_frame, dup_data_frame = get_data()
    data_process(all_policies_dedup_data_frame, dup_data_frame, colors_dict)
